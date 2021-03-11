# -*- coding: utf-8 -*-


"""
Created on Thu Mar  11 9:56:00 2021

Balance Study to extract name/balance from monthly financials

@author: johnmillstead
"""

import pandas as pd
import openpyxl as op
from openpyxl import load_workbook
from pathlib import Path

# Excel File
src_file = src_file = Path.cwd() / 'test_balance.xlsx'

df = load_workbook(filename=src_file, read_only=True)

balance_df = pd.DataFrame(df.sheetnames)
